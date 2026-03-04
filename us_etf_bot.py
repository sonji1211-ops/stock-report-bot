import os, pandas as pd, asyncio, datetime
import yfinance as yf
from telegram import Bot
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

TOKEN = "8574978661:AAF5SXIgfpJlnAfN5ccSk0tJek_uSlCMBBo"
CHAT_ID = "8564327930"

# [전종목 리스트 그대로 유지]
ASSET_NAMES = {
    'KS11': '코스피 지수', 'KQ11': '코스닥 지수', 
    'USD/KRW': '달러/원 환율', 'JPY/KRW': '엔/원 환율', 
    'EUR/KRW': '유로/원 환율', 'CNY/KRW': '위안/원 환율', 
    '069500.KS': 'KODEX 200', '252670.KS': 'KODEX 200선물인버스2X', '305720.KS': 'KODEX 2차전지산업',
    '462330.KS': 'KODEX AI반도체핵심공정', '122630.KS': 'KODEX 레버리지',
    'BTC-KRW': '비트코인', 'ETH-KRW': '이더리움', 'XRP-KRW': '리플(XRP)', 
    'SOL-KRW': '솔라나(SOL)', 'USDT-KRW': '테더(USDT)',
    'QQQ': '나스닥100', 'TQQQ': '나스닥100(3배)', 'SQQQ': '나스닥100인버스(3배)', 'QLD': '나스닥100(2배)',
    'SPY': 'S&P500', 'IVV': 'S&P500(iShares)', 'VOO': 'S&P500(Vanguard)', 'SSO': 'S&P500(2배)', 'Upro': 'S&P500(3배)',
    'DIA': '다우존스', 'IWM': '러셀2000', 'SOXX': '필라델피아반도체', 'SOXL': '반도체강세(3배)', 'SOXS': '반도체약세(3배)',
    'SMH': '반도체ETF(VanEck)', 'NVDL': '엔비디아(2배)', 'TSLL': '테슬라(2배)', 'CONL': '코인베이스(2배)',
    'SCHD': '슈드(배당성장)', 'JEPI': '제피(고배당)', 'ARKK': '아크혁신(캐시우드)',
    'TLT': '미국채20년(장기채)', 'TMF': '장기채강세(3배)', 'TMV': '장기채약세(3배)',
    'XLF': '금융섹터', 'XLV': '헬스케어섹터', 'XLE': 'energy섹터', 'XLK': '기술주섹터', 
    'XLY': '임의소비재', 'XLP': '필수소비재', 'GDX': '금광업', 'GLD': '금선물',
    'VNQ': '리츠(부동산)', 'BITO': '비트코인ETF', 'FNGU': '빅테크플러스(3배)', 'BULZ': '빅테크성장(3배)',
    'VTI': '미국전체주식', 'VXUS': '미국외전세계', 'VT': '전세계주식',
    'GC=F': '금 선물', 'SI=F': '은 선물'
}

def format_num(val):
    """소수점 아래가 0이면 정수로, 아니면 소수점 2자리까지 표시"""
    if val == int(val): return int(val)
    return round(val, 2)

async def fetch_asset_data(symbol):
    try:
        yf_symbol = symbol
        if symbol == 'KS11': yf_symbol = '^KS11'
        elif symbol == 'KQ11': yf_symbol = '^KQ11'
        elif symbol == 'USD/KRW': yf_symbol = 'KRW=X'
        elif symbol == 'JPY/KRW': yf_symbol = 'JPYKRW=X'
        elif symbol == 'EUR/KRW': yf_symbol = 'EURKRW=X'
        elif symbol == 'CNY/KRW': yf_symbol = 'CNYKRW=X'
        elif symbol.isdigit(): yf_symbol = symbol + ".KS"

        ticker_obj = yf.Ticker(yf_symbol)
        df = ticker_obj.history(period="7d")
        
        # 실제 값이 있는 데이터만 남기기
        df = df.dropna(subset=['Close'])
        if df.empty or len(df) < 2: return None

        # 전일 종가 대비 정확한 계산 (마지막 두 거래일 비교)
        last_c = float(df['Close'].iloc[-1])
        prev_c = float(df['Close'].iloc[-2])
        ratio = ((last_c - prev_c) / prev_c) * 100
            
        return {'티커': symbol, '항목명': ASSET_NAMES.get(symbol, symbol), '현재가': format_num(last_c), '등락률': format_num(ratio)}
    except: return None

async def main():
    bot = Bot(token=TOKEN)
    now = datetime.datetime.utcnow() + datetime.timedelta(hours=9)
    
    print(f"📡 전일 대비 데이터 수집 중...")
    results = []
    for s in ASSET_NAMES.keys():
        res = await fetch_asset_data(s)
        if res: results.append(res)
        await asyncio.sleep(0.05)

    if not results: return

    df = pd.DataFrame(results)
    file_name = f"{now.strftime('%m%d')}_종합_리포트.xlsx"
    
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df.rename(columns={'등락률':'등락률(%)'}).to_excel(writer, sheet_name='현황', index=False)
        ws = writer.sheets['현황']
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_fill = PatternFill(start_color='444444', end_color='444444', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12

        for r in range(1, ws.max_row + 1):
            for c in range(1, 5):
                cell = ws.cell(r, c)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                if r == 1: cell.fill, cell.font = header_fill, white_font
                
            if r > 1:
                ticker = str(ws.cell(r, 1).value)
                ratio_val = float(ws.cell(r, 4).value or 0)
                
                # [숫자 포맷팅] 소수점이 0이면 정수형태로 보이게 엑셀 서식 적용
                # 기본 서식은 소수점 최대 2자리까지
                ws.cell(r, 3).number_format = '#,##0.##'
                ws.cell(r, 4).number_format = '0.##"%"'

                # 화폐 기호 추가
                is_won = any(x in ticker for x in ['-KRW', '/KRW', 'KS11', 'KQ11']) or ticker.replace('.KS','').isdigit()
                if is_won:
                    ws.cell(r, 3).number_format = '"₩"#,##0.##'
                else:
                    ws.cell(r, 3).number_format = '"$"#,##0.##'
                
                # ±5% 이상 노란색 강조
                if abs(ratio_val) >= 5:
                    for c in range(1, 5): ws.cell(r, c).fill = yellow_fill

    async with bot:
        await bot.send_document(CHAT_ID, open(file_name, 'rb'), 
                               caption=f"🌍 종합 리포트 ({now.strftime('%Y-%m-%d')})\n✅ 전일 대비 계산 및 소수점 0 제거 완료")
    if os.path.exists(file_name): os.remove(file_name)

if __name__ == "__main__":
    asyncio.run(main())
