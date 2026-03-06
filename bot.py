import os, pandas as pd, asyncio, datetime, requests, time
from telegram import Bot
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from urllib.parse import unquote

# [1. 설정 정보]
TELEGRAM_TOKEN = "8574978661:AAF5SXIgfpJlnAfN5ccSk0tJek_uSlCMBBo"
CHAT_ID = "8564327930"
RAW_KEY = "3e937f2b0780c88e27c6f4cb99d5b58e69cc71cef898809e7aacb2bcabe1b438"
SERVICE_KEY = unquote(RAW_KEY)

def get_official_data(target_date=None):
    url = 'http://apis.data.go.kr/1160100/service/GetStockSecuritiesInfoService/getStockPriceInfo'
    params = {'serviceKey': SERVICE_KEY, 'numOfRows': '4000', 'resultType': 'json'}
    if target_date: params['basDt'] = target_date.replace('-', '')
    try:
        res = requests.get(url, params=params, timeout=30).json()
        items = res['response']['body']['items'].get('item', [])
        if not items:
            params['serviceKey'] = RAW_KEY
            res = requests.get(url, params=params, timeout=30).json()
            items = res['response']['body']['items'].get('item', [])
        return pd.DataFrame(items)
    except: return pd.DataFrame()

async def main():
    bot = Bot(token=TELEGRAM_TOKEN)
    now = datetime.datetime.utcnow() + datetime.timedelta(hours=9)
    day_of_week = now.weekday() # 0:월, 1:화, ..., 5:토, 6:일

    df = pd.DataFrame()
    analysis_info = ""
    mode_name = ""

    # [요일별 로직 분기]
    # 1. 일요일 (6): 주간평균리포트 (월~금 데이터 요약)
    if day_of_week == 6: 
        mode_name = "주간평균리포트"
        weekly_dfs = []
        # 지난주 월(6일전)~금(2일전) 데이터를 수집
        for i in range(2, 7):
            target_dt = (now - datetime.timedelta(days=i)).strftime('%Y%m%d')
            df_day = get_official_data(target_dt)
            if not df_day.empty and len(df_day) > 500:
                for col in ['mkp', 'clpr', 'lopr', 'hipr', 'fltRt', 'trqu']:
                    df_day[col] = pd.to_numeric(df_day[col], errors='coerce').fillna(0)
                weekly_dfs.append(df_day)
        
        if weekly_dfs:
            full_df = pd.concat(weekly_dfs)
            df = full_df.groupby(['itmsNm', 'srtnCd', 'mrktCtg']).agg({
                'mkp': 'first', 'clpr': 'last', 'lopr': 'min', 'hipr': 'max', 'fltRt': 'mean', 'trqu': 'mean'
            }).reset_index()
            df.columns = ['종목명', '종목코드', '시장', '시가', '종가', '저가', '고가', '등락률(%)', '거래량']
            analysis_info = "기간: 지난주 월요일 ~ 금요일 (평균)"
        else:
            print("주간 데이터를 찾을 수 없습니다.")
            return

    # 2. 화(1) ~ 토(5): 일일마감리포트 (어제 데이터 기준)
    elif 1 <= day_of_week <= 5:
        mode_name = "일일마감리포트"
        found = False
        # 어제(1일전)부터 혹시 몰라 3일전까지 탐색 (공공데이터 지연 대비)
        for i in range(1, 4):
            search_date = (now - datetime.timedelta(days=i)).strftime('%Y%m%d')
            raw = get_official_data(search_date)
            if not raw.empty and len(raw) > 500:
                df = pd.DataFrame()
                df['종목코드'] = raw['srtnCd']
                df['종목명'] = raw['itmsNm']
                df['시가'] = pd.to_numeric(raw['mkp']).fillna(0).astype(int)
                df['종가'] = pd.to_numeric(raw['clpr']).fillna(0).astype(int)
                df['저가'] = pd.to_numeric(raw['lopr']).fillna(0).astype(int)
                df['고가'] = pd.to_numeric(raw['hipr']).fillna(0).astype(int)
                df['등락률(%)'] = pd.to_numeric(raw['fltRt']).fillna(0).astype(float)
                df['거래량'] = pd.to_numeric(raw['trqu']).fillna(0).astype(int)
                df['시장'] = raw['mrktCtg']
                analysis_info = f"마감 기준일: {search_date}"
                found = True
                break
        if not found:
            async with bot:
                await bot.send_message(CHAT_ID, "⚠️ 국장 데이터가 아직 업데이트되지 않았습니다.")
            return
    
    # 3. 월요일(0) 또는 기타: 리포트 미발송
    else:
        print("오늘은 리포트를 생성하지 않는 요일입니다.")
        return

    if df.empty: return

    # [엑셀 생성 및 스타일링 - 지수님 요청 디자인]
    target_cols = ['종목코드', '종목명', '시가', '종가', '저가', '고가', '등락률(%)', '거래량']
    file_name = f"{now.strftime('%m%d')}_{mode_name}.xlsx"
    
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    font_white_bold = Font(color="FFFFFF", bold=True)
    colors = {
        'red': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        'orange': PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
        'yellow': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    }
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        for market in ['KOSPI', 'KOSDAQ']:
            for is_up in [True, False]:
                sheet_label = f"{market}_{'상승' if is_up else '하락'}"
                m_cond = df['시장'].str.contains(market)
                r_cond = (df['등락률(%)'] >= 5) if is_up else (df['등락률(%)'] <= -5)
                sub_df = df[m_cond & r_cond].copy().sort_values('등락률(%)', ascending=not is_up)
                sub_df = sub_df[target_cols]
                
                sub_df.to_excel(writer, sheet_name=sheet_label, index=False)
                ws = writer.sheets[sheet_label]

                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 20

                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8), 1):
                    for c_idx, cell in enumerate(row, 1):
                        cell.border = border
                        cell.alignment = center_align
                        if r_idx == 1:
                            cell.fill, cell.font = header_fill, font_white_bold
                        else:
                            if c_idx in [3, 4, 5, 6, 8]: cell.number_format = '#,##0'
                            if c_idx == 7:
                                cell.number_format = '0.00'
                                val = abs(float(cell.value or 0))
                                name_cell = ws.cell(row=r_idx, column=2)
                                if val >= 25: 
                                    name_cell.fill, name_cell.font = colors['red'], font_white_bold
                                elif val >= 20: name_cell.fill = colors['orange']
                                elif val >= 10: name_cell.fill = colors['yellow']

    total_up = len(df[df['등락률(%)'] >= 5])
    total_down = len(df[df['등락률(%)'] <= -5])
    
    msg = (f"📊 모드: {mode_name}\n🔍 {analysis_info}\n\n"
           f"📈 상승(5%↑): {total_up}개\n📉 하락(5%↓): {total_down}개\n"
           f"💡 🟡10%↑ 🟠20%↑ 🔴25%↑")

    async with bot:
        await bot.send_document(CHAT_ID, open(file_name, 'rb'), caption=msg)
    if os.path.exists(file_name): os.remove(file_name)

if __name__ == "__main__":
    asyncio.run(main())
