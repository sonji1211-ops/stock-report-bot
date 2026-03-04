import os, pandas as pd, asyncio, datetime, requests, time
from telegram import Bot
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import FinanceDataReader as fdr

# [1. 설정 정보]
TOKEN = "8574978661:AAF5SXIgfpJlnAfN5ccSk0tJek_uSlCMBBo"
CHAT_ID = "8564327930"

# [2. 주요 종목 한글 매핑]
KR_NAMES = {
    'AAPL': '애플', 'MSFT': '마이크로소프트', 'NVDA': '엔비디아', 'TSLA': '테슬라', 
    'AMZN': '아마존', 'META': '메타', 'GOOGL': '알파벳A', 'AVGO': '브로드컴',
    'NFLX': '넷플릭스', 'AMD': 'AMD', 'MU': '마이크론', 'QCOM': '퀄컴',
    'ORCL': '오라클', 'COST': '코스트코', 'ADBE': '어도비', 'INTC': '인텔',
    'BRK-B': '버크셔헤서웨이', 'V': '비자', 'MA': '마스터카드', 'JPM': 'JP모건'
}

async def fetch_stock_safe(row, start_dt, end_dt):
    """차단 방지를 위한 개별 종목 수집 (오류 시 무시)"""
    symbol = row['Symbol']
    try:
        # 데이터 수집
        df = fdr.DataReader(symbol, start_dt, end_dt)
        if df.empty or len(df) < 2: return None
        
        last_close = float(df.iloc[-1]['Close'])
        prev_close = float(df.iloc[-2]['Close'])
        ratio = round(((last_close - prev_close) / prev_close) * 100, 2)
        
        return {
            '티커': symbol,
            '종목명': KR_NAMES.get(symbol, row.get('Name', symbol)),
            '종가': last_close,
            '등락률(%)': ratio,
            '산업': row.get('Industry', '-'),
            '기준일': df.index[-1].strftime('%Y-%m-%d')
        }
    except:
        return None

async def main():
    bot = Bot(token=TOKEN)
    now = datetime.datetime.utcnow() + datetime.timedelta(hours=9)
    
    print("📡 미국 시장(NASDAQ+NYSE) 상위 1,000개 리스트 구성 중...")
    try:
        # 1. 미국 전체 시장 상위 종목 수집 (중복 제거 및 상위 1,000개)
        df_nasdaq = fdr.StockListing('NASDAQ')
        df_nyse = fdr.StockListing('NYSE')
        df_base = pd.concat([df_nasdaq, df_nyse]).drop_duplicates('Symbol')
        df_target = df_base.head(1000) # 시총 상위 1,000개
        
        start_dt = (now - datetime.timedelta(days=10)).strftime('%Y-%m-%d')
        end_dt = now.strftime('%Y-%m-%d')
        
        results = []
        chunk_size = 20 # 20개씩 끊어서 요청
        
        print(f"🚀 분석 시작 (총 {len(df_target)}개 종목)... 약 10분 소요 예정")
        for i in range(0, len(df_target), chunk_size):
            chunk = df_target.iloc[i:i+chunk_size]
            tasks = [fetch_stock_safe(row, start_dt, end_dt) for _, row in chunk.iterrows()]
            chunk_results = await asyncio.gather(*tasks)
            results.extend([r for r in chunk_results if r is not None])
            
            # 진행상황 출력 및 휴식 (중요!)
            print(f"⏳ 진행 중: {min(i+chunk_size, len(df_target))}/{len(df_target)} 완료")
            await asyncio.sleep(2.0) # 1,000개 수집 시 차단 방지를 위해 2초 휴식

        df_final = pd.DataFrame(results)
        if df_final.empty: return

        # 2. 엑셀 파일 생성
        file_name = f"{now.strftime('%m%d')}_미국장_1000_리포트.xlsx"
        target_cols = ['티커', '종목명', '종가', '등락률(%)', '산업', '기준일']
        
        h_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
        f_white = Font(color="FFFFFF", bold=True)
        colors = {
            'red': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'orange': PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
            'yellow': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        }
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            for trend in ['상승', '하락']:
                cond = (df_final['등락률(%)'] >= 5) if trend == '상승' else (df_final['등락률(%)'] <= -5)
                sub = df_final[cond].copy().sort_values('등락률(%)', ascending=(trend == '하락'))
                sub = sub[target_cols]
                
                sheet_name = f"미국_{trend}"
                sub.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                # [너비 설정] 종목명 30 고정
                ws.column_dimensions['A'].width = 12 # 티커
                ws.column_dimensions['B'].width = 30 # 종목명
                ws.column_dimensions['C'].width = 15 # 종가
                ws.column_dimensions['D'].width = 15 # 등락률
                ws.column_dimensions['E'].width = 40 # 산업
                ws.column_dimensions['F'].width = 15 # 기준일

                # [스타일링]
                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6), 1):
                    for c_idx, cell in enumerate(row, 1):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        if r_idx == 1:
                            cell.fill, cell.font = h_fill, f_white
                        else:
                            if c_idx == 3: cell.number_format = '#,##0.00'
                            if c_idx == 4: # 등락률 소수점 및 B열 종목명 색상 강조
                                cell.number_format = '0.00'
                                rv = abs(float(cell.value or 0))
                                name_cell = ws.cell(row=r_idx, column=2)
                                if rv >= 25: name_cell.fill, name_cell.font = colors['red'], f_white
                                elif rv >= 20: name_cell.fill = colors['orange']
                                elif rv >= 10: name_cell.fill = colors['yellow']

        # 3. 텔레그램 전송
        up_count = len(df_final[df_final['등락률(%)'] >= 5])
        down_count = len(df_final[df_final['등락률(%)'] <= -5])
        
        async with bot:
            msg = (f"🇺🇸 미국 시장(NASDAQ/NYSE) 상위 1000\n"
                   f"📈 상승(5%↑): {up_count}개\n"
                   f"📉 하락(5%↓): {down_count}개\n"
                   f"💡 1,000개 전수 조사 완료 (B열 색상 강조)")
            await bot.send_document(CHAT_ID, open(file_name, 'rb'), caption=msg)
        
        if os.path.exists(file_name): os.remove(file_name)
        print("✅ 미국장 1,000개 리포트 전송 성공!")

    except Exception as e:
        print(f"🚨 오류: {e}")

if __name__ == "__main__":
    asyncio.run(main())
