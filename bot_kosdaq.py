import os, pandas as pd, asyncio, time
from yahooquery import Ticker
from datetime import datetime, timedelta
from telegram import Bot
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

TOKEN = "8574978661:AAF5SXIgfpJlnAfN5ccSk0tJek_uSlCMBBo"
CHAT_ID = "8564327930"

def get_kosdaq_data_final():
    print("📡 [야후 쿼리 엔진] 코스닥 전수조사 가동...")
    
    # [코스닥 핵심 종목 리스트] 시총 상위 및 변동성 종목 위주 (약 80개)
    codes = [
        '086520', '091990', '247540', '293490', '066970', '028300', '068270', '035900', '214150', '058470',
        '035760', '263750', '112040', '036830', '039030', '041510', '051910', '078340', '041960', '034230',
        '086900', '056190', '048410', '067160', '036490', '278280', '145020', '240810', '067310', '214450',
        '084370', '053800', '032500', '092040', '000250', '253450', '064550', '036120', '046890', '196170',
        '023160', '089010', '200130', '060250', '041190', '051370', '040300', '033640', '065350', '042000',
        '035600', '063170', '121600', '036200', '046120', '060720', '038500', '043200', '032190', '054670',
        '049480', '054920', '036710', '085660', '108490', '065680', '042700', '035080', '054630', '040420',
        '043370', '052670', '069330', '033320', '036540', '141080', '086450', '039200', '131970', '215600'
    ]
    # 코스닥은 접미사가 .KQ 입니다.
    tickers = [c + ".KQ" for c in codes]
    
    try:
        t = Ticker(tickers, asynchronous=True)
        data = t.price
        
        all_stocks = []
        for symbol, info in data.items():
            if isinstance(info, dict) and 'regularMarketPrice' in info:
                ratio = info.get('regularMarketChangePercent', 0) * 100
                all_stocks.append({
                    'Name': symbol.split('.')[0],
                    'Open': info.get('regularMarketOpen', 0),
                    'Close': info.get('regularMarketPrice', 0),
                    'Low': info.get('regularMarketDayLow', 0),
                    'High': info.get('regularMarketDayHigh', 0),
                    'Ratio': float(ratio),
                    'Volume': info.get('regularMarketVolume', 0)
                })
        
        print(f"✅ 코스닥 {len(all_stocks)}개 종목 확보 성공!")
        return pd.DataFrame(all_stocks)
    except Exception as e:
        print(f"❌ 코스닥 엔진 구동 실패: {e}")
        return pd.DataFrame()

async def main():
    bot = Bot(token=TOKEN)
    now = datetime.utcnow() + timedelta(hours=9)
    df = get_kosdaq_data_final()
    
    if df.empty:
        print("🚨 코스닥 데이터 수집 실패.")
        return

    r_type = "주간평균" if now.weekday() == 6 else "일일"
    file_name = f"{now.strftime('%m%d')}_KOSDAQ_{r_type}.xlsx"
    
    # 5% 필터링
    up_df = df[df['Ratio'] >= 5.0].sort_values('Ratio', ascending=False)
    down_df = df[df['Ratio'] <= -5.0].sort_values('Ratio', ascending=True)

    # 디자인 설정 (코스피와 동일한 지수님 요청 양식)
    h_map = {'Name':'종목명','Open':'시가','Close':'종가','Low':'저가','High':'고가','Ratio':'등락률(%)','Volume':'거래량'}
    red_f, ora_f, yel_f = PatternFill("solid", "FF0000"), PatternFill("solid", "FFCC00"), PatternFill("solid", "FFFF00")
    head_f, white_font = PatternFill("solid", "444444"), Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        for s_name, d in {'코스닥_상승': up_df, '코스닥_하락': down_df}.items():
            tmp = d.rename(columns=h_map) if not d.empty else pd.DataFrame([['조건 만족 종목 없음']+['']*6], columns=list(h_map.values()))
            tmp.to_excel(writer, sheet_name=s_name, index=False)
            ws = writer.sheets[s_name]
            
            for cell in ws[1]:
                cell.fill, cell.font, cell.alignment, cell.border = head_f, white_font, center_align, thin_border

            for r in range(2, ws.max_row + 1):
                try:
                    ratio_val = abs(float(ws.cell(r, 6).value))
                    if ratio_val >= 28: ws.cell(r, 1).fill, ws.cell(r, 1).font = red_f, white_font
                    elif ratio_val >= 20: ws.cell(r, 1).fill = ora_f
                    elif ratio_val >= 10: ws.cell(r, 1).fill = yel_f
                except: pass

                for c in range(1, 8):
                    ws.cell(r, c).alignment, ws.cell(r, c).border = center_align, thin_border
                    if c in [2, 3, 4, 5, 7]: ws.cell(r, c).number_format = '#,##0'
                    if c == 6: ws.cell(r, c).number_format = '0.00'
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15

    msg = f"📅 {now.strftime('%m-%d')} *[KOSDAQ {r_type}]*\n📈 분석대상: {len(df)}개\n🚀 상승(5%↑): {len(up_df)}개 / 하락(5%↓): {len(down_df)}개"
    try:
        with open(file_name, 'rb') as f:
            await bot.send_document(CHAT_ID, document=f, caption=msg, parse_mode="Markdown")
        print("🚀 코스닥 전송 성공!")
    except Exception as e:
        print(f"❌ 코스닥 전송 실패: {e}")
    finally:
        if os.path.exists(file_name): os.remove(file_name)

if __name__ == "__main__":
    asyncio.run(main())
