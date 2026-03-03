import os
import pandas as pd
import requests
import re
import io
import time
import random
from datetime import datetime, timedelta
import asyncio
from bs4 import BeautifulSoup
from telegram import Bot
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# [설정] 텔레그램 정보
TOKEN = "8574978661:AAF5SXIgfpJlnAfN5ccSk0tJek_uSlCMBBo"
CHAT_ID = "8564327930"

def fetch_naver_page(url, headers):
    """네이버 서버 차단을 우회하기 위한 직접 HTML 요청"""
    try:
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code == 200:
            return response.text
    except:
        return None
    return None

async def get_kospi_market_scan():
    """요구사항 1: 코스피 전수조사 (BeautifulSoup 정밀 파싱)"""
    fields = "field=quant&field=open&field=high&field=low&field=frate"
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
    ]
    
    all_stocks = []
    # market_code 0 = KOSPI
    headers = {'User-Agent': random.choice(user_agents), 'Referer': 'https://finance.naver.com/'}
    
    # 1. 마지막 페이지 확인
    init_text = fetch_naver_page(f"https://finance.naver.com/sise/sise_market_sum.naver?sosok=0", headers)
    if not init_text: return pd.DataFrame()
    
    last_page_match = re.findall(r'page=(\d+)', init_text)
    last_page = int(max(map(int, last_page_match))) if last_page_match else 1
    
    print(f"📡 KOSPI 전수조사 시작 ({last_page}페이지)...")

    for page in range(1, last_page + 1):
        url = f"https://finance.naver.com/sise/sise_market_sum.naver?sosok=0&{fields}&page={page}"
        success = False
        
        for attempt in range(3):
            headers['User-Agent'] = random.choice(user_agents)
            html = fetch_naver_page(url, headers)
            
            if html and "종목명" in html:
                soup = BeautifulSoup(html, 'html.parser')
                table = soup.find('table', {'class': 'type_2'})
                if table:
                    df_list = pd.read_html(io.StringIO(str(table)))
                    if df_list:
                        df = df_list[0].dropna(subset=['종목명']).copy()
                        df.columns = [c.strip() for c in df.columns]
                        
                        if '시가' in df.columns:
                            cols = ['등락률', '현재가', '시가', '고가', '저가', '거래량']
                            for col in cols:
                                if col in df.columns:
                                    df[col] = df[col].astype(str).str.replace('%','').str.replace(',','').str.replace('+','').replace('nan', '0')
                                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                            
                            for _, row in df.iterrows():
                                all_stocks.append({
                                    'Name': str(row['종목명']), 'Open': int(row['시가']), 'Close': int(row['현재가']),
                                    'Low': int(row['저가']), 'High': int(row['고가']), 'Ratio': float(row['등락률']),
                                    'Volume': int(row['거래량'])
                                })
                            success = True
                            break
            time.sleep(random.uniform(0.5, 1.0)) # 차단 회피를 위해 지연시간 소폭 증가
        
        if page % 10 == 0: print(f"✅ KOSPI {page}p 완료...")
            
    return pd.DataFrame(all_stocks)

async def send_kospi_report():
    bot = Bot(token=TOKEN)
    now = datetime.utcnow() + timedelta(hours=9)
    
    try:
        df_final = await get_kospi_market_scan()
        if df_final.empty:
            print("❌ KOSPI 데이터 수집 실패"); return

        report_type = "주간평균" if now.weekday() == 6 else "일일"
        file_name = f"{now.strftime('%m%d')}_KOSPI_{report_type}.xlsx"

        # 필터링 로직 (상승/하락 5% 기준)
        def get_sub(is_up):
            temp = df_final[df_final['Volume'] > 0].copy()
            cond = (temp['Ratio'] >= 5.0) if is_up else (temp['Ratio'] <= -5.0)
            return temp[cond].sort_values('Ratio', ascending=not is_up)

        sheets = {
            '코스피_상승': get_sub(True),
            '코스피_하락': get_sub(False)
        }

        # 디자인 디테일 (색상, 콤마, 소수점, 테두리)
        h_map = {'Name':'종목명','Open':'시가','Close':'종가','Low':'저가','High':'고가','Ratio':'등락률(%)','Volume':'거래량'}
        red, orange, yellow = PatternFill("solid", "FF0000"), PatternFill("solid", "FFCC00"), PatternFill("solid", "FFFF00")
        header_fill, white_font = PatternFill("solid", "444444"), Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            for s_name, data in sheets.items():
                data.rename(columns=h_map).to_excel(writer, sheet_name=s_name, index=False)
                ws = writer.sheets[s_name]
                
                # 헤더 스타일
                for cell in ws[1]:
                    cell.fill, cell.font, cell.alignment = header_fill, white_font, Alignment(horizontal='center')
                
                # 본문 서식
                for r in range(2, ws.max_row + 1):
                    val = abs(float(ws.cell(r, 6).value or 0))
                    # 종목명 색상 강조 규칙 (A열)
                    if val >= 28: ws.cell(r, 1).fill, ws.cell(r, 1).font = red, white_font
                    elif val >= 20: ws.cell(r, 1).fill = orange
                    elif val >= 10: ws.cell(r, 1).fill = yellow
                    
                    for c in range(1, 8):
                        ws.cell(r, c).alignment, ws.cell(r, c).border = Alignment(horizontal='center'), border
                        if c in [2, 3, 4, 5, 7]: ws.cell(r, c).number_format = '#,##0'
                        if c == 6: ws.cell(r, c).number_format = '0.00'
                
                ws.column_dimensions['A'].width = 18
                for i in range(2, 8): ws.column_dimensions[chr(64+i)].width = 13

        # 통계 메시지 구성
        up_cnt = len(sheets['코스피_상승'])
        down_cnt = len(sheets['코스피_하락'])
        msg = (f"📅 {now.strftime('%m-%d')} *[KOSPI {report_type}]*\n"
               f"📊 분석: 코스피 전 종목 ({len(df_final)}개) 전수조사\n"
               f"📈 상승(5%↑): {up_cnt} / 📉 하락(5%↓): {down_cnt}\n"
               f"💡 🔴28%↑, 🟠20%↑, 🟡10%↑ (모든 하락주 포함)")
        
        with open(file_name, 'rb') as f:
            await bot.send_document(CHAT_ID, document=f, caption=msg, parse_mode="Markdown")
        os.remove(file_name)
        print("✅ KOSPI 전용 리포트 전송 성공!")

    except Exception as e:
        print(f"❌ KOSPI 오류: {e}")

if __name__ == "__main__":
    asyncio.run(send_kospi_report())
